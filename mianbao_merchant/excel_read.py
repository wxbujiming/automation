"""
Excel 文件读取模块

提供从 Excel 文件中读取 API 接口列表的功能，支持解析参数、请求头和请求体。
支持多平台（多个 Sheet 页）读取。
"""

import os
import json
import openpyxl
from config import EXCEL_PATH
from loging import *


def _parse_params(params_str):
    """
    解析参数字符串为字典
    
    支持格式：{pageSize=10,pageNum=1} 或 pageSize=10,pageNum=1
    
    Args:
        params_str: 参数字符串
        
    Returns:
        dict: 解析后的参数字典
        
    Examples:
        >>> _parse_params("{pageSize=10,pageNum=1}")
        {'pageSize': '10', 'pageNum': '1'}
        
        >>> _parse_params("key1=value1,key2=value2")
        {'key1': 'value1', 'key2': 'value2'}
    """
    params = {}
    if not params_str or not isinstance(params_str, str):
        return params
    
    try:
        # 去掉前后的大括号和空格
        clean_str = params_str.strip().strip("{} ").strip()
        if not clean_str:
            return params
        
        # 用逗号分割参数对
        param_pairs = clean_str.split(",")
        for pair in param_pairs:
            pair = pair.strip()
            if "=" in pair:
                key, value = pair.split("=", 1)  # 只分割第一个等号
                params[key.strip()] = value.strip()
    except Exception as e:
        logger.error(f"⚠️  解析参数失败：{params_str} - {type(e).__name__}: {e}")
    
    return params


def _parse_headers(headers_str):
    """
    解析请求头字符串为字典
    
    支持格式：Content-Type:application/json,Authorization:Bearer xxx
    
    Args:
        headers_str: 请求头字符串
        
    Returns:
        dict: 解析后的请求头字典
        
    Examples:
        >>> _parse_headers("Content-Type:application/json,Authorization:Bearer token123")
        {'Content-Type': 'application/json', 'Authorization': 'Bearer token123'}
    """
    headers = {}
    if not headers_str or not isinstance(headers_str, str):
        return headers
    
    try:
        for h in headers_str.split(","):
            h = h.strip()
            if ":" in h:
                k, v = h.split(":", 1)
                headers[k.strip()] = v.strip()
    except Exception as e:
        logger.error(f"⚠️  解析请求头失败：{headers_str} - {type(e).__name__}: {e}")
    
    return headers


def _parse_body(body_str):
    """
    解析请求体 JSON 字符串
    
    Args:
        body_str: JSON 格式的字符串
        
    Returns:
        dict/list/None: 解析后的对象，失败返回 None
        
    Examples:
        >>> _parse_body('{"key": "value"}')
        {'key': 'value'}
        
        >>> _parse_body("invalid json")
        None
    """
    if not body_str or not isinstance(body_str, str):
        return None
    
    try:
        return json.loads(body_str.strip())
    except json.JSONDecodeError as e:
        logger.error(f"⚠️  解析请求体失败：{body_str[:50]}... - {type(e).__name__}: {e}")
        return None




def _parse_assertion(assertion_str):
    """
    解析断言字符串为字典
    
    支持 JSON 格式，用于定义响应验证规则
    例如：{"code": 0, "data.title": "存在关键字"} 或 {"status": "success"}
    
    Args:
        assertion_str: 断言 JSON 字符串
        
    Returns:
        dict/None: 解析后的断言字典，失败返回 None
        
    Examples:
        >>> _parse_assertion('{"code": 0, "message": "success"}')
        {'code': 0, 'message': 'success'}
        
        >>> _parse_assertion('{"data.user.name": "admin"}')
        {'data.user.name': 'admin'}
    """
    if not assertion_str or not isinstance(assertion_str, str):
        return None
    
    try:
        return json.loads(assertion_str.strip())
    except json.JSONDecodeError as e:
        logger.error(f"⚠️  解析断言失败：{assertion_str[:50]}... - {type(e).__name__}: {e}")
        return None


def read_excel_api_list(excel_path=EXCEL_PATH):
    """
    从 Excel 文件读取所有平台的 API 接口列表
    
    自动读取所有工作表中的数据，按平台标识（Sheet 名称）分组返回。
    
    Args:
        excel_path: Excel 文件路径，默认使用配置中的 EXCEL_PATH
        
    Returns:
        dict: {
            "merchant": [api_info1, api_info2, ...],  # 运营平台接口列表
            "platform": [api_info1, api_info2, ...]   # 管理平台接口列表
        }
        每个 api_info 包含以下字段：
            - api_id: 接口编号
            - api_name: 接口名称
            - method: 请求方式
            - path: 接口路径
            - params: 参数字典
            - headers: 请求头字典
            - body: 请求体对象（dict/list）
            - assertion: 断言字典（用于验证响应）
            - group: 所属平台（Sheet 名称）
        
    Raises:
        FileNotFoundError: Excel 文件不存在
        Exception: 其他读取错误
        
    Example:
        >>> api_dict = read_excel_api_list()
        >>> len(api_dict["merchant"])
        10
        >>> len(api_dict["platform"])
        5
    """
    api_dict = {}
    wb = None
    
    try:
        # 验证文件是否存在
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Excel 文件不存在：{excel_path}")
        
        # 加载工作簿和工作表（read_only 模式提高性能）
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        
        # 遍历所有 Sheet 页
        for sheet_name in wb.sheetnames:
            logger.info(f"📖 正在读取工作表：{sheet_name}")
            ws = wb[sheet_name]
            api_list = []
            
            # 读取数据（第 2 行开始，跳过表头）
            for row in ws.iter_rows(min_row=2, values_only=True):
                # 对应表头：接口编号 (0)、接口名称 (1)、请求方式 (2)、接口路径 (3)、
                # Params 参数 (4)、请求头 (5)、请求体 (6)、断言 (7)
                if len(row) < 8:
                    continue
                    
                api_id, api_name, method, path, params_str, headers_str, body_str, assertion_str = row[:8]
                
                # 跳过空行（path 为空）
                if not path or (isinstance(path, str) and not path.strip()):
                    continue
                
                # 解析各个字段
                params = _parse_params(params_str)
                headers = _parse_headers(headers_str)
                body = _parse_body(body_str)
                assertion = _parse_assertion(assertion_str)
                
                # 组装接口信息（添加 group 字段标识平台）
                api_info = {
                    "api_id": api_id,
                    "api_name": api_name,
                    "method": method,
                    "path": path.strip() if isinstance(path, str) else path,
                    "params": params,
                    "headers": headers,
                    "body": body,
                    "assertion": assertion,
                    "group": sheet_name  # 使用 Sheet 名称作为分组标识
                }
                api_list.append(api_info)
                logger.debug(f"组装接口信息{api_info}")
            
            if api_list:
                api_dict[sheet_name] = api_list
                logger.info(f"✅ 工作表 '{sheet_name}' 成功读取 {len(api_list)} 个接口")
            else:
                logger.warning(f"⚠️  工作表 '{sheet_name}' 为空，已跳过")
        
        total_count = sum(len(apis) for apis in api_dict.values())
        logger.info(f"📊 总计：从 {len(api_dict)} 个工作表读取 {total_count} 个接口")
        
    except FileNotFoundError as e:
        logger.error(f"❌ 文件未找到：{e}")
        raise
    except Exception as e:
        logger.error(f"❌ 读取 Excel 失败：{type(e).__name__} - {e}")
        raise
    finally:
        # 确保关闭工作簿
        if wb is not None:
            wb.close()
    
    return api_dict


if __name__ == "__main__":
    # 测试代码
    logger.info("=" * 60)
    logger.info("Excel 读取模块测试（多平台支持）")
    logger.info("=" * 60)
    
    try:
        # 读取 Excel 文件（自动读取所有 Sheet）
        api_dict = read_excel_api_list(EXCEL_PATH)
        
        # 打印结果
        logger.info(f"\n📋 共读取 {len(api_dict)} 个平台\\n")
        
        for platform, api_list in api_dict.items():
            logger.info(f"{'='*60}")
            logger.info(f"平台：{platform} | 接口数：{len(api_list)}")
            logger.info(f"{'='*60}")
            
            for i, api in enumerate(api_list[:3], 1):  # 每个平台只显示前 3 个
                logger.info(f"  {i}. {api['api_name']}")
                logger.info(f"     路径：{api['path']}")
                logger.info(f"     方式：{api['method']}")
                if api['params']:
                    logger.info(f"     参数：{api['params']}")
                if api['headers']:
                    logger.info(f"     请求头：{api['headers']}")
                if api['body']:
                    logger.info(f"     请求体：{api['body']}")
                if api['assertion']:
                    logger.info(f"     断言：{api['assertion']}")
                logger.info()
            
            if len(api_list) > 3:
                logger.info(f"  ... 还有 {len(api_list) - 3} 个接口\\n")
            
    except Exception as e:
        logger.info(f"❌ 测试失败：{type(e).__name__} - {e}")
